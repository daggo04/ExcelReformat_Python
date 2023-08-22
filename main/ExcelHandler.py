import openpyxl
from openpyxl.utils import cell as cell_utils
from openpyxl.utils import quote_sheetname, absolute_coordinate
from openpyxl.styles import NamedStyle, Font, Border, PatternFill, Alignment
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

class ExcelHandler:
    def __init__(self, source_path: Path, template_path: Path = None):
        self.source = openpyxl.load_workbook(source_path)
        
        if template_path:
            self.template = openpyxl.load_workbook(template_path)
            self.output = openpyxl.Workbook()
            for sheetname in self.template.sheetnames:
                self.output.create_sheet(sheetname)
                self._copy_worksheet(self.template[sheetname], self.output[sheetname])
                
            # Copy defined names
            self._copy_defined_names()
            # Remove the default sheet created by openpyxl
            default_sheet = self.output.active
            if default_sheet:
                self.output.remove(default_sheet)
        else:
            self.template = None

    def save_output_workbook(self, output_path: Path):
        self.output.save(output_path)

    def copy_rows(self, src_sheet: float, dst_sheet: float, start_row: float, end_row: float):
        source_sheet = self.source.worksheets[int(src_sheet)]
        target_sheet = self.output.worksheets[int(dst_sheet)]

        for i in range(int(start_row), int(end_row) + 1):
            source_row = source_sheet[i]
            target_row = target_sheet[i]
            for source_cell, target_cell in zip(source_row, target_row):
                self._copy_cell(source_cell, target_cell)

    def copy_column(self, src_sheet: float, src_col: float, dst_sheet: float, dst_col: float, start_row: float):
        src_col_idx = int(src_col)
        dst_col_idx = int(dst_col)
        start_row_idx = int(start_row)

        source_sheet = self.source.worksheets[int(src_sheet)]
        target_sheet = self.output.worksheets[int(dst_sheet)]

        for row in source_sheet.iter_rows(min_row=start_row_idx, max_col=src_col_idx + 1, min_col=src_col_idx):
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=dst_col_idx + 1, value=cell.value)
                self._copy_cell(cell, target_cell)

    def copy_split_row(self, src_sheet: float, dst_sheet: float, start_row: float, col_map: dict, include_headers: bool = False, header_col: float = 0):
        source_sheet = self.source.worksheets[int(src_sheet)]
        target_sheet = self.output.worksheets[int(dst_sheet)]
        
        last_source_row = source_sheet.max_row
        target_row_idx = int(start_row+1)

        # Convert string keys to integer keys
        col_map = {int(float(key)): int(float(value)) for key, value in col_map.items()}

        inverted_col_map = self._invert_col_map(col_map) #Invert the col_map to get a map of source columns to target columns

        splits, key_to_split = self._determine_splits_and_key(inverted_col_map) #Find the key with the most values and the number of splits

        for i in range(int(start_row+1), last_source_row + 1):
            source_row = source_sheet[i]
            #Break if the next 10 rows are empty
            if self._are_next_rows_empty(source_sheet, i, 10): 
                break
            
             # If the row is empty, continue to the next iteration
            if not source_row or not any(cell.value for cell in source_row):
                continue
            
            
            for j in range(splits): #For each split
                target_row = target_sheet[target_row_idx]
                target_row_idx += 1
                
                # Copy cells
                for key, values in inverted_col_map.items(): #For each source column
                    if key == key_to_split: #If this is the column to split on
                        source_cell = source_row[values[min(j, len(values) - 1)]] #Get the value from the column corresponding to the current split
                    else:
                        source_cell = source_row[values[0]] #Get the value from the first column

                    target_cell = target_row[int(key)] #Get the target cell
                    self._copy_cell(source_cell, target_cell) #Copy the value

                    # Copy headers
                    if include_headers: #If headers should be included 
                        header_cell = source_sheet.cell(1, column= source_cell.column) #Get the header cell
                        target_header_cell = target_row[int(header_col)] #Get the target header cell
                        self._copy_cell(header_cell, target_header_cell)

    def set_column_value(self, sheet: float, col: float, value: str, start_row: float, end_row: float = -1):
        target_sheet = self.output.worksheets[int(sheet)]
        dst_col_idx = int(col)

        # If end_row is not provided, find the last populated row in all columns and pick the longest two
        if end_row == -1:
            populated_columns = []

            # Scan through the first row and identify populated columns
            for column in range(1, target_sheet.max_column + 1):
                if target_sheet.cell(row=2, column=column).value:
                    populated_columns.append(column)

            lengths = []
            for column in populated_columns:
                for row in range(target_sheet.max_row, 0, -1):
                    if target_sheet.cell(row=row, column=column).value:
                        lengths.append(row)
                        break

            # If there are populated columns, get the longest
            if lengths:
                lengths.sort(reverse=True)
                end_row = lengths[0]
                if len(lengths) > 1:
                    end_row = max(lengths[:1])

            # If no populated columns, default to max_row
            else:
                end_row = target_sheet.max_row

        for row in range(start_row, end_row + 1):
            cell = target_sheet.cell(row=row, column=dst_col_idx)
            cell.value = value
        
    # Helper methods
    @staticmethod
    def _copy_cell(source_cell, target_cell):
        # Copy cell value
        target_cell.value = source_cell.value
        
        
        # Check if the source cell value is a string
        if isinstance(source_cell.value, str) and target_cell.row != 1:
            return
        elif source_cell.value is None:
            return

        # If you want to copy style, don't directly pass the proxy style
        # Instead, create a new font/border/alignment/fill from the source's attributes
        if source_cell.has_style:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
            
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom,
            )
            
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                textRotation=source_cell.alignment.textRotation,
                wrapText=source_cell.alignment.wrapText,
                shrinkToFit=source_cell.alignment.shrinkToFit,
                indent=source_cell.alignment.indent,
                relativeIndent=source_cell.alignment.relativeIndent,
                justifyLastLine=source_cell.alignment.justifyLastLine,
            )
            
            target_cell.fill = PatternFill(
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color,
                patternType=source_cell.fill.patternType
            )

            # Copy the number_format (date formatting)
            target_cell.number_format = source_cell.number_format

    @staticmethod
    def _copy_worksheet(source_sheet: Worksheet, destination_sheet: Worksheet):
        for i, row in enumerate(source_sheet.iter_rows()):
            for j, cell in enumerate(row):
                target_cell = destination_sheet.cell(row=i + 1, column=j + 1)
                ExcelHandler._copy_cell(cell, target_cell)
                
    # Helper methods for copy_split_row
    @staticmethod
    def _invert_col_map(col_map: dict) -> dict:
        inverted = {}
        for k, v in col_map.items():
            inverted[v] = inverted.get(v, [])
            inverted[v].append(k)
        return inverted
    
    @staticmethod
    def _determine_splits_and_key(inverted_col_map: dict) -> tuple:
        splits = 1
        key_to_split = 0
        for key, values in inverted_col_map.items():
            if len(values) > splits:
                splits = len(values)
                key_to_split = key
        return splits, key_to_split

    @staticmethod
    def _are_next_rows_empty(sheet, start_row, num_rows) -> bool:
        for i in range(start_row, start_row + num_rows):
            if sheet[i]:
                return False
        return True
    
    def _copy_defined_names(self):
        # Copy globally defined names
        for name, defn in self.template.defined_names.items():
            destinations = defn.destinations
            for title, coord in destinations:
                output_sheet = self.output.get_sheet_by_name(title)
                
                # Check if the name already contains a sheet reference
                if "!" in coord:
                    ref = coord
                else:
                    ref = f"{quote_sheetname(output_sheet.title)}!{absolute_coordinate(coord)}"
                
                defn_new = DefinedName(name, attr_text=ref)
                self.output.defined_names[name] = defn_new

        # Copy worksheet-specific defined names
        for sheet in self.template.worksheets:
            for name, defn in sheet.defined_names.items():
                output_sheet = self.output.get_sheet_by_name(sheet.title)
                
                # Check if the name already contains a sheet reference
                if "!" in defn.attr_text:
                    ref = defn.attr_text
                else:
                    ref = f"{quote_sheetname(output_sheet.title)}!{absolute_coordinate(defn.attr_text)}"
                
                defn_new = DefinedName(name, attr_text=ref)
                output_sheet.defined_names.add(defn_new)

            